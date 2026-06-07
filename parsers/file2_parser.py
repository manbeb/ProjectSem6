import openpyxl
import re
import os
from typing import Dict, Optional


def parse_file2(filepath: str, department: str = "Не указана") -> Optional[Dict]:
    wb = openpyxl.load_workbook(filepath, data_only=True)

    # Ищем лист с итогами
    sheet = None
    for name in wb.sheetnames:
        if 'ИТОГ' in name.upper():
            sheet = wb[name]
            break

    if not sheet:
        wb.close()
        return None

    # 1. Извлекаем ФИО и Должность из первых строк
    fio = "Неизвестный"
    position = "Не указана"

    for row_idx in range(1, 15):
        cell_val = str(sheet.cell(row=row_idx, column=1).value or " ")
        if any(pos in cell_val for pos in ['Заведующий', 'Доцент', 'Профессор', 'Старший', 'Ассистент']):
            match_fio = re.match(r'^([А-Яа-яЁё\.\s\-]+?)(?:\s{2,}|,)', cell_val)
            if match_fio:
                fio = match_fio.group(1).strip()

            match_pos = re.search(r'(Заведующий кафедрой|Доцент|Профессор|Старший преподаватель|Ассистент)', cell_val)
            if match_pos:
                position = match_pos.group(1).strip()
            break

    # 2. Парсим таблицу итогов (суммируем часы)
    totals = {}
    fact_hours = None

    for row_idx in range(1, sheet.max_row + 1):
        cell_b = str(sheet.cell(row=row_idx, column=2).value or " ")
        cell_c = sheet.cell(row=row_idx, column=3).value

        if 'Фактическое кол-во часов' in cell_b or 'Фактическое кол-во' in cell_b:
            fact_hours = float(cell_c) if cell_c is not None else None

        if cell_c is not None:
            if 'Учебная нагрузка' in cell_b:
                totals['Учебная'] = float(cell_c)
            elif 'Неконтактная' in cell_b:
                totals['Неконтактная'] = float(cell_c)
            elif 'Метод.' in cell_b:
                totals['Метод'] = float(cell_c)
            elif 'Электр.' in cell_b:
                totals['Электр'] = float(cell_c)
            elif 'Научная' in cell_b:
                totals['Научная'] = float(cell_c)
            elif 'Орг.' in cell_b:
                totals['Орг'] = float(cell_c)
            elif 'Повыш.' in cell_b:
                totals['Повыш'] = float(cell_c)
            elif 'Поручения' in cell_b:
                totals['Поручения'] = float(cell_c)

    plan_total = sum(v for v in totals.values() if v is not None)
    wb.close()

    return {
        'ФИО': fio,
        'Должность': position,
        'Кафедра': department,  # <-- Добавили кафедру из имени папки
        'План_Из_Файла2': plan_total if plan_total > 0 else totals.get('Учебная', 0),
        'Факт_Из_Файла2': fact_hours,
        'Источник': os.path.basename(filepath)
    }