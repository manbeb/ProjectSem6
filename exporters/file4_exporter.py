import pandas as pd
import os
import re
import openpyxl
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter
from config import YELLOW_FILL, LIGHT_GRAY_FILL


def _apply_formatting(filepath: str, df: pd.DataFrame):
    """Применяет стилизацию и авто-ширину столбцов к Excel-файлу."""
    wb = openpyxl.load_workbook(filepath)
    ws = wb.active
    df = df.reset_index(drop=True)

    # 1. Стилизация шапки (голубой больше не используем, используем светло-серый)
    for col_idx, col_name in enumerate(df.columns, start=1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = Font(color="000000", bold=True)
        cell.alignment = Alignment(wrap_text=True, vertical="center")

        # Расчёт авто-ширины
        max_length = len(str(col_name))
        for val in df.iloc[:, col_idx - 1]:
            val_str = str(val) if pd.notna(val) else ""
            if len(val_str) > max_length:
                max_length = len(val_str)

        new_width = min(max_length + 2, 60)
        ws.column_dimensions[get_column_letter(col_idx)].width = new_width

    # 2. Подсветка строк по статусу
    for row_idx, row in df.iterrows():
        status = str(row.get('Статус', ''))
        excel_row = row_idx + 2

        # Приоритет: сначала проверяем отсутствие, потом расхождение
        if 'Отсутствует' in status:
            fill_color = LIGHT_GRAY_FILL
        elif 'Расхождение' in status:
            fill_color = YELLOW_FILL
        else:
            fill_color = None

        if fill_color:
            for col in range(1, len(df.columns) + 1):
                ws.cell(row=excel_row, column=col).fill = fill_color

    wb.save(filepath)
    wb.close()


def export_reports(result_df: pd.DataFrame, general_dir: str, departments_dir: str) -> dict:
    os.makedirs(general_dir, exist_ok=True)
    os.makedirs(departments_dir, exist_ok=True)

    paths = {'general': '', 'departments': {}}

    # 1. Общий отчёт
    general_path = os.path.join(general_dir, "Отчёт_общий.xlsx")
    result_df.to_excel(general_path, index=False, sheet_name="Сравнение")
    _apply_formatting(general_path, result_df)
    paths['general'] = general_path

    # 2. Отчёты по кафедрам с АГРЕГАЦИЕЙ
    df_clean = result_df.copy()
    df_clean['Кафедра'] = df_clean['Кафедра'].fillna("Не указана").astype(str).str.strip()

    numeric_cols = [
        'План (ИС ВВГУ)', 'Факт (Отчёт кафедры)', 'Разница (План - Факт)',
        'Учебная нагрузка', 'Неконтактная работа', 'Метод. работа',
        'Электр. обучение', 'Научная работа', 'Орг. работа',
        'Повыш. квалификации', 'Поручения рук. напр.'
    ]

    # Умная функция агрегации статуса
    def aggregate_status(statuses):
        statuses_list = [str(v) for v in statuses]
        # Если хотя бы одна запись отсутствует в ИС ВВГУ, флаг должен быть максимальным
        if any('Отсутствует' in s for s in statuses_list):
            return '❌ Отсутствует в ИС ВВГУ'
        # Иначе ищем любое расхождение
        if any('Расхождение' in s for s in statuses_list):
            for s in statuses_list:
                if 'Расхождение' in s:
                    return s
        return '✓ Совпадает'

    grouped = df_clean.groupby('Кафедра')
    for dept, group_df in grouped:
        clean_dept = re.sub(r'[\\/*?:"<>|]', "_", dept)
        if not clean_dept:
            clean_dept = "Без_кафедры"

        agg_dict = {col: 'sum' for col in numeric_cols if col in group_df.columns}

        # Применяем умную агрегацию статуса
        agg_dict['Статус'] = aggregate_status
        agg_dict['Файлы-источники'] = lambda x: ', '.join(
            sorted(list(set(y.strip() for val in x for y in str(val).split(',')))))

        # АГРЕГИРУЕМ данные СТРОГО по ФИО ППС
        agg_df = group_df.groupby(['ФИО ППС'], as_index=False).agg(agg_dict)

        for col in numeric_cols:
            if col in agg_df.columns:
                agg_df[col] = agg_df[col].round(2)

        # Удаляем колонку 'Кафедра' и 'Должность' для отчёта по кафедре
        agg_df = agg_df.drop(columns=['Кафедра', 'Должность'], errors='ignore')

        dept_path = os.path.join(departments_dir, f"Отчёт_{clean_dept}.xlsx")
        agg_df.to_excel(dept_path, index=False, sheet_name="Сравнение")
        _apply_formatting(dept_path, agg_df)
        paths['departments'][dept] = dept_path

    return paths