import pandas as pd
from typing import Tuple


class WorkloadComparator:
    TOLERANCE_HOURS = 5  # Допуск ±5 часов

    def __init__(self, file1_df: pd.DataFrame, mapping_df: pd.DataFrame = None):
        self.is_data = file1_df
        self.mapping = mapping_df
        self.results = []

    def normalize_name(self, name: str) -> str:
        """Нормализация ФИО для надёжного сопоставления"""
        # Убираем лишние пробелы, приводим к верхнему регистру
        return ' '.join(name.strip().upper().split())

    def find_match(self, name: str, department: str) -> Tuple[bool, pd.Series]:
        """
        Поиск соответствия в данных ИС ВВГУ
        Возвращает (найдено, строка_данных)
        """
        norm_name = self.normalize_name(name)

        # Прямое совпадение ФИО + Кафедра
        mask = (
                (self.is_data['ФИО_очищенное'].apply(self.normalize_name) == norm_name) &
                (self.is_data['Кафедра'] == department)
        )

        matches = self.is_data[mask]
        if not matches.empty:
            return True, matches.iloc[0]

        # Если не найдено — пробуем только по ФИО (для случаев смены кафедры)
        mask_name_only = self.is_data['ФИО_очищенное'].apply(self.normalize_name) == norm_name
        matches_name = self.is_data[mask_name_only]

        if not matches_name.empty:
            return True, matches_name.iloc[0]

        return False, None

    def compare_record(self, file2_record: dict, department: str) -> dict:
        """
        Сравнение одной записи из Файла 2 с данными Файла 1
        """
        result = {
            'ФИО': file2_record['ФИО'],
            'Кафедра': department,
            'План_ИС_ВВГУ': None,
            'Факт_из_плана': file2_record.get('факт_часов'),
            'Разница': None,
            'Статус': '✓',
            'Подсветка': False
        }

        found, is_row = self.find_match(file2_record['ФИО'], department)

        if found:
            result['План_ИС_ВВГУ'] = is_row['План']

            if result['Факт_из_плана'] is not None:
                diff = result['Факт_из_плана'] - result['План_ИС_ВВГУ']
                result['Разница'] = diff

                # Проверка допуска ±5 часов
                if abs(diff) > self.TOLERANCE_HOURS:
                    result['Статус'] = '⚠ Расхождение'
                    result['Подсветка'] = True  # Жёлтый цвет
                else:
                    result['Статус'] = '✓ В допуске'
            else:
                result['Статус'] = 'ℹ Нет данных "Факт"'
        else:
            # Запись есть в Файле 2, но отсутствует в Файле 1
            result['Статус'] = '❌ Не найдено в ИС ВВГУ'
            result['Подсветка'] = True  # Жёлтый цвет

        return result


from openpyxl import load_workbook
from openpyxl.styles import PatternFill

YELLOW_FILL = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")


def export_to_file4(results: list, output_path: str):
    """
    Создаёт Файл 4: Excel с результатами сравнения и условным форматированием
    """
    df = pd.DataFrame(results)

    # Сохраняем базовый файл
    df.to_excel(output_path, index=False, sheet_name='Сравнение')

    # Применяем форматирование через openpyxl
    wb = load_workbook(output_path)
    ws = wb.active

    # Находим колонку "Подсветка" (предполагаем, что это последняя)
    highlight_col_idx = len(df.columns)  # 1-based index для openpyxl

    # Проходим по строкам данных (начиная со 2-й, т.к. 1-я — заголовок)
    for row_idx, record in enumerate(results, start=2):
        if record.get('Подсветка'):
            # Закрашиваем всю строку в жёлтый
            for col_idx in range(1, len(df.columns) + 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.fill = YELLOW_FILL

    # Добавляем легенду
    legend_row = len(results) + 3
    ws[f'A{legend_row}'] = "Легенда:"
    ws[f'B{legend_row}'] = "🟡 Жёлтый = расхождение >5 часов или запись не найдена в ИС"

    wb.save(output_path)
    wb.close()